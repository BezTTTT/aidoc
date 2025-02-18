from flask import (
    Blueprint, redirect, render_template, request, session, url_for, g, flash, current_app, send_from_directory,jsonify,Flask,Response
)
import io
import pandas as pd
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from aidoc.utils import *
from aidoc.db import get_db
from aidoc.auth import login_required, admin_only, specialist_only, role_validation, reload_user_profile

# 'webapp' blueprint manages Diagnosis and Record systems, including report and admin managment system
bp = Blueprint('webapp', __name__)

# Flask views

# region aidoc logo
@bp.route('/favicon.ico') 
def favicon(): 
    return send_from_directory(os.path.join(current_app.root_path, 'static', 'icons'), 'favicon.ico', mimetype='image/vnd.microsoft.icon')

# region quick_confirm
@bp.route('/quick_confirm/<role>/<int:img_id>', methods=('POST', ))
@login_required
@role_validation
def quick_confirm(role, img_id):
    if role=='specialist' or role=='admin':
        ai_result = request.args.get('ai_result', type=int)
        if ai_result==0:
            diagnostic_code = 'NORMAL'
        elif ai_result==1:
            diagnostic_code = 'OPMD'
        else:
            diagnostic_code = 'OSCC'
        db, cursor = get_db()
        sql = "UPDATE submission_record SET dentist_id=%s, dentist_feedback_code=%s, dentist_feedback_date=%s, updated_at=%s WHERE id=%s"
        val = ( session["user_id"], diagnostic_code, datetime.now(), datetime.now(), img_id)
        cursor.execute(sql, val)
    return redirect(url_for('webapp.record', role=role, page=session['current_record_page']))

# region followup_request
@bp.route('/followup_request/<role>/<int:img_id>', methods=('POST', ))
@login_required
@role_validation
def followup_request(role, img_id):

    db, cursor = get_db()
    sql = "SELECT * FROM followup_request WHERE submission_id=%s"
    val = ( img_id,)
    cursor.execute(sql, val)
    data = cursor.fetchall()
    if not data:
        sql = "INSERT INTO followup_request (submission_id, followup_requester, followup_request_status) VALUES (%s, %s, %s)"
        val = ( img_id, session["user_id"], 'Initiated')
        cursor.execute(sql, val)
    else:
        sql = "DELETE FROM followup_request WHERE submission_id=%s"
        val = ( img_id, )
        cursor.execute(sql, val)
    return redirect(url_for('webapp.record', role=role, page=session['current_record_page']))

# region retrain_request
@bp.route('/retrain_request/<role>/<int:img_id>', methods=('POST', ))
@login_required
@role_validation
def retrain_request(role, img_id):
    db, cursor = get_db()
    sql = "SELECT * FROM retrain_request WHERE submission_id=%s"
    val = ( img_id,)
    cursor.execute(sql, val)
    data = cursor.fetchall()
    if not data: # check is the list is empty
        sql = "INSERT INTO retrain_request (submission_id, retrain_requester, retrain_request_status) VALUES (%s, %s, %s)"
        val = ( img_id, session["user_id"], 'Requested')
        cursor.execute(sql, val)
    else:
        sql = "DELETE FROM retrain_request WHERE submission_id=%s"
        val = ( img_id, )
        cursor.execute(sql, val)
    return redirect(url_for('webapp.record', role=role, page=session['current_record_page']))

# region diagnosis
@bp.route('/diagnosis/<role>/<int:img_id>', methods=('GET', 'POST'))
@login_required
@role_validation
def diagnosis(role, img_id):

    # In the case of returning from the register_later system, clear the variable
    if 'register_later' in session:
        session.pop('register_later', None)

    if request.method=='POST':
        db, cursor = get_db()
        if request.args.get('special_request')=='true':
            sql = "UPDATE submission_record SET special_request=%s WHERE id=%s"
            val = (1, img_id)
            cursor.execute(sql, val)

        if (role=='dentist' or (role=='admin' and request.args.get('channel')=='DENTIST')) and request.form.get('dentist_action'):
            dentist_action_code = request.form.get('dentist_action')
            if(dentist_action_code == 'ai_agreement'):
                dentist_feedback_code = request.form.get('agree_option')
                sql = "UPDATE submission_record SET dentist_id=%s, dentist_feedback_code=%s, dentist_feedback_date=%s WHERE id=%s"
                val = (session["user_id"], dentist_feedback_code, datetime.now(), img_id)
                cursor.execute(sql, val)
            if(dentist_action_code == 'additional_feedback'):
                dentist_feedback_location = request.form.get('lesion_location')
                dentist_feedback_lesion = request.form.get('lesion_type')
                sql = "UPDATE submission_record SET dentist_id=%s, dentist_feedback_lesion=%s, dentist_feedback_location=%s, dentist_feedback_date=%s WHERE id=%s"
                val = (session["user_id"], dentist_feedback_lesion, dentist_feedback_location, datetime.now(), img_id)
                cursor.execute(sql, val)
            if(dentist_action_code == 'comment'):
                dentist_feedback_comment = request.form.get('dentist_comment')
                sql = "UPDATE submission_record SET dentist_id=%s, dentist_feedback_comment=%s, dentist_feedback_date=%s WHERE id=%s"
                val = (session["user_id"], dentist_feedback_comment, datetime.now(), img_id)
                cursor.execute(sql, val)
    
        if (role=='specialist' or (role=='admin' and request.args.get('channel')!='DENTIST')) and request.args.get('specialist_feedback')=='true':
            dentist_feedback_code = request.form.get('dt_comment_option')
            dentist_feedback_lesion = None
            dentist_feedback_location = None
            dentist_feedback_comment = ''
            if dentist_feedback_code=='BAD_IMG':
                dentist_feedback_comment = request.form.get('BadImgCommentSelectOptions')
            elif dentist_feedback_code=='OPMD' or dentist_feedback_code=='OSCC':
                dentist_feedback_lesion = request.form.get('LesionTypeSelection')
                dentist_feedback_location = request.form.get('LesionLocationSelection')
            elif dentist_feedback_code=='OTHER':
                dentist_feedback_comment = request.form.get('OtherCommentTextarea', '')
            elif dentist_feedback_code=='BENIGN':
                dentist_feedback_comment = request.form.get('BenignCommentSelectOptions')
            sql = '''UPDATE submission_record SET
                        dentist_id=%s,
                        dentist_feedback_code=%s,
                        dentist_feedback_comment=%s,
                        dentist_feedback_lesion=%s,
                        dentist_feedback_location=%s,
                        dentist_feedback_date=%s,
                        updated_at=%s
                    WHERE id=%s'''
            val = (session["user_id"],
                   dentist_feedback_code,
                   dentist_feedback_comment,
                   dentist_feedback_lesion,
                   dentist_feedback_location,
                   datetime.now(),
                   datetime.now(),
                   img_id)
            cursor.execute(sql, val)

        if (role=='specialist' or (role=='admin' and request.args.get('channel')!='DENTIST')) and request.args.get('case_report')=='true':
            case_report = request.form.get('case_report')
            sql = '''UPDATE submission_record SET
                        dentist_id=%s,
                        case_report=%s,
                        dentist_feedback_date=%s,
                        updated_at=%s
                    WHERE id=%s'''
            val = (session["user_id"],
                   case_report,
                   datetime.now(),
                   datetime.now(),
                   img_id)
            cursor.execute(sql, val)

    dentistFeedbackRequest = request.args.get('dentistFeedbackRequest', None) # Request from patient_record --> patient_diagnosis

    db, cursor = get_db()
    if role=='specialist' or (role=='admin' and request.args.get('channel')!='DENTIST'):
        sql = '''SELECT submission_record.id AS img_id, case_id, channel, fname, special_request,
                    patient_id, patient.name AS patient_name, patient.surname AS patient_surname, patient_national_id as saved_patient_national_id, patient.national_id AS db_patient_national_id, patient.birthdate,
                    patient.sex, patient.job_position, patient.email, patient.phone AS patient_phone, patient.address, patient.province,
                    sender_id, sender.name AS sender_name, sender.surname AS sender_surname, sender.hospital AS sender_hospital, sender.province AS sender_province,
                    sender.phone AS sender_phone_db, sender_phone,
                    dentist_id, dentist.name AS dentist_name, dentist.surname AS dentist_surname,
                    dentist_feedback_code, dentist_feedback_comment, dentist_feedback_lesion, dentist_feedback_location, dentist_feedback_date, case_report,
                    location_district, location_amphoe, location_province, location_zipcode,
                    ai_prediction, ai_scores, lesion_ai_version, quality_ai_prediction, quality_ai_version, ai_updated_at, submission_record.created_at
                FROM submission_record
                INNER JOIN patient_case_id ON submission_record.id = patient_case_id.id
                LEFT JOIN user AS sender ON submission_record.sender_id = sender.id
                LEFT JOIN user AS patient ON submission_record.patient_id = patient.id
                LEFT JOIN user AS dentist ON submission_record.dentist_id = dentist.id
                WHERE submission_record.id=%s
                LIMIT 1'''
    elif role=='osm':
        sql = '''SELECT submission_record.id AS img_id, case_id, channel, fname, special_request, sender_id, patient_id, sender_phone, sender.phone AS osm_phone, 
                    patient.name AS patient_name, patient.surname AS patient_surname, patient_national_id as saved_patient_national_id, patient.national_id AS db_patient_national_id, patient.birthdate,
                    patient.sex, patient.job_position, patient.email, patient.phone AS patient_phone, patient.address, patient.province,
                    location_district, location_amphoe, location_province, location_zipcode,
                    ai_prediction, ai_scores, lesion_ai_version, quality_ai_prediction, quality_ai_version, ai_updated_at, submission_record.created_at
                FROM submission_record
                INNER JOIN patient_case_id ON submission_record.id=patient_case_id.id
                LEFT JOIN user AS sender ON submission_record.sender_phone = sender.phone
                LEFT JOIN user AS patient ON submission_record.patient_id = patient.id
                WHERE submission_record.id=%s
                LIMIT 1'''
    elif role=='dentist' or (role=='admin' and request.args.get('channel')=='DENTIST'):
        sql = '''SELECT submission_record.id AS img_id, fname, sender_id,
                    dentist_feedback_code, dentist_feedback_comment, dentist_feedback_lesion, dentist_feedback_location,
                    ai_prediction, ai_scores, lesion_ai_version, quality_ai_prediction, quality_ai_version, ai_updated_at
                FROM submission_record
                WHERE id=%s
                LIMIT 1'''
    else: # Patient system
        if dentistFeedbackRequest and dentistFeedbackRequest=='true':
            sql = '''SELECT submission_record.id AS img_id, case_id, channel, fname, special_request, sender_id, patient_id, sender_phone, 
                    dentist_feedback_code, dentist_feedback_comment, dentist_feedback_date,
                    ai_prediction, ai_scores, lesion_ai_version, quality_ai_prediction, quality_ai_version, ai_updated_at, submission_record.created_at
                FROM submission_record
                INNER JOIN patient_case_id ON submission_record.id=patient_case_id.id
                WHERE submission_record.id=%s
                LIMIT 1'''
        else: # Patient system without dentist feedback (dentist feedback is pending)
            sql = '''SELECT submission_record.id AS img_id, case_id, channel, fname, special_request, sender_id, patient_id, sender_phone, 
                    ai_prediction, ai_scores, lesion_ai_version, quality_ai_prediction, quality_ai_version, ai_updated_at, submission_record.created_at
                FROM submission_record
                INNER JOIN patient_case_id ON submission_record.id=patient_case_id.id
                WHERE submission_record.id=%s
                LIMIT 1'''
        
    val = (img_id, )
    cursor.execute(sql, val)
    data = cursor.fetchone()

    # Authorization check
    if data is None:
        return render_template('unauthorized_access.html', error_msg='ไม่พบข้อมูลที่ร้องขอ Data Not Found')
    elif \
        (role=='patient' and (session['user_id']!=data['patient_id'])) or \
        (role=='osm' and session['user_id']!=data['sender_id'] and data['sender_phone']!=data['osm_phone']) or \
        (role=='dentist' and session['user_id']!=data['sender_id']):
            return render_template('unauthorized_access.html', error_msg='คุณไม่มีสิทธิ์เข้าถึงข้อมูล Unauthorized Access')

    # Further process the data
    
    if role=='admin':
        data['channel'] = request.args.get('channel')

    if 'channel' in data and data['channel']=='PATIENT':
        data['owner_id'] = data['patient_id']
    else:
        data['owner_id'] = data['sender_id']

    # AI Results Processing
    data['ai_scores'] = json.loads(data['ai_scores'])
    if data['lesion_ai_version']:
        data['lesion_ai_version_check'] = data['lesion_ai_version'] == current_app.config['AI_LESION_VER']
    if data['quality_ai_version']:
        data['quality_ai_version_check'] = data['quality_ai_version'] == current_app.config['AI_QUALITY_VER']
    if data['ai_updated_at']:
        data['ai_updated_at_thai_datetime'] = format_thai_datetime(data['ai_updated_at'])
    else:
        data['ai_updated_at_thai_datetime'] = 'ไม่มีผลลัพธ์ (กด Recompute ใหม่)'

    if role=='osm' or role=='specialist' or (role=='admin' and request.args.get('channel')!='DENTIST'):
        # Sender info processing
        if role=='specialist' or (role=='admin' and request.args.get('channel')!='DENTIST'):
            if data['patient_id']!=data['sender_id']:
                if 'sender_name' in data and data['sender_name'] is not None:
                    data['sender_description'] = f"{data['sender_name']} {data['sender_surname']} (ผู้นำส่งข้อมูล, เบอร์โทรติดต่อ: {data['sender_phone_db']})"
                else:
                    data['sender_description'] = f"ผู้นำส่งข้อมูล เบอร์โทรติดต่อ: {data['sender_phone']} (ยังไม่ได้ลงทะเบียน)"
            else:
                data['sender_description'] = f"{data['sender_name']} {data['sender_surname']} (ผู้ป่วยนำส่งรูปด้วยตัวเอง)"
        else: # osm
            data['sender_description'] = f"{g.user['name']} {g.user['surname']} (โทรศัพท์: {g.user['phone']})"
            data['sender_hospital'] = g.user['hospital']
            data['sender_province'] = g.user['province']

        # Datetime format processing
        data['thai_datetime'] = format_thai_datetime(data['created_at'])

        # Location format processing
        if data['location_district']:
            data['location_text'] = "ตำบล"+data['location_district']+" อำเภอ"+data['location_amphoe']+" จังหวัด"+data['location_province']+" " +str(data['location_zipcode'])
        else:
            data['location_text'] = "จังหวัด"+data['location_province']

        # Age and sex processing
        if 'birthdate' in data and data['birthdate'] is not None:
            data['patient_age'] = calculate_age(data['birthdate'])
            if data['sex']=='M':
                data['sex']='ชาย'
            elif data['sex']=='F':
                data['sex']='หญิง'

    # For patient_dignosis page for the case that has the dentist feedback
    if dentistFeedbackRequest:
        data['dentistFeedbackRequest'] = dentistFeedbackRequest
        if data['ai_prediction']==0 and data['dentist_feedback_code']=='NORMAL':
            data['dentistCommentAgreeCode'] = 'TN'
        elif data['ai_prediction']==0 and (data['dentist_feedback_code']=='OPMD' or data['dentist_feedback_code']=='OSCC'):
            data['dentistCommentAgreeCode'] = 'FN'
        elif data['ai_prediction']!=0 and (data['dentist_feedback_code']=='OPMD' or data['dentist_feedback_code']=='OSCC'):
            data['dentistCommentAgreeCode'] = 'TP'
        elif data['ai_prediction']!=0 and data['dentist_feedback_code']=='NORMAL':
            data['dentistCommentAgreeCode'] = 'FP'
        else:
            data['dentistCommentAgreeCode'] = 'Error'
            if data['dentist_feedback_comment'] == 'NON_STANDARD':
                data['dentistComment'] = 'มุมมองไม่ได้มาตรฐาน'
            elif data['dentist_feedback_comment'] == 'BLUR':
                data['dentistComment'] = 'ภาพเบลอ ไม่ชัด'
            elif data['dentist_feedback_comment'] == 'DARK':
                data['dentistComment'] = 'ภาพช่องปากมืดเกินไป ขอเปิดแฟลชด้วย'
            elif data['dentist_feedback_comment'] == 'SMALL':
                data['dentistComment'] = 'ช่องปากเล็กเกินไป ขอนำกล้องเข้าใกล้ปากมากกว่านี้'
    else:
        data['dentistFeedbackRequest'] = 'false'

    dentist_diagnosis_map = {'NORMAL': 'ยืนยันว่าไม่พบรอยโรค (Normal)',
                                'BENIGN': 'น่าจะไม่มีรอยโรค (Benign)',
                                'OPMD': 'น่าจะมีรอยโรคที่คล้ายกันกับ OPMD',
                                'OSCC': 'น่าจะมีรอยโรคที่คล้ายกันกับ OSCC',
                                'BAD_IMG': 'ภาพถ่ายที่ส่งมายังไม่ได้มาตรฐาน ทำให้วินิจฉัยไม่ได้',
                                'OTHER': 'ความเห็น/คำวินิจฉัย อื่น ๆ ที่ต้องการแจ้งกลับให้ผู้ป่วย'
                            }
    bad_image_map = {'NON_STANDARD': 'มุมมองไม่ได้มาตรฐาน',
                     'BLUR': 'ภาพเบลอ ไม่ชัด',
                     'DARK': 'ภาพช่องปากมืดเกินไป ขอเปิดแฟลชด้วย',
                     'SMALL': 'ช่องปากเล็กเกินไป ขอนำกล้องเข้าใกล้ปากมากกว่านี้'
                    }
    lesion_location_map = {1: 'ริมฝีปากบนและล่าง (Lip)',
                           2: 'กระพุ้งแก้มด้านขวาด้านซ้าย (Buccal mucosa)',
                           3: 'เหงือกบนและล่าง (Gingiva)',
                           4: 'เหงือกด้านหลังฟันกรามล่าง (Retromolar area)',
                           5: 'เพดานแข็ง (Hard palate)',
                           6: 'เพดานอ่อน (Soft palate)',
                           7: 'ลิ้นด้านบนและด้านข้าง (Dorsal and lateral tongue)',
                           8: 'ใต้ลิ้น (Ventral tongue)',
                           9: 'พื้นปาก (Floor of mouth)'
                           }
    lesion_type_map = {1: 'รอยโรคสีขาว',
                       2: 'รอยโรคสีแดง',
                       3: 'รอยโรคสีขาวปนแดง',
                       4: 'รอยแผลถลอก',
                       5: 'ลักษณะเป็นก้อน'
                       }
    
    benign_option = {'NORMAL': 'ปกติ ไม่ใช่รอยโรค',
                    'RECHECK': 'ควรตรวจเพิ่มเติม',  
                    'OBSERVE': 'ติดตามอาการเพิ่มเติม'}


    maps = {'dentist_diagnosis_map': dentist_diagnosis_map,
            'bad_image_map': bad_image_map,
            'lesion_location_map': lesion_location_map,
            'lesion_type_map': lesion_type_map,
            'benign_option': benign_option}
    return render_template(role+'_diagnosis.html', data=data, maps=maps)

# region record
@bp.route('/record/<role>', methods=('GET', 'POST'))
@login_required
@role_validation
def record(role): 

    # There is no pagination for patient's record page (a special case)
    if role=='patient':
        data = record_patient()
        return render_template("patient_record.html", data=data)

    # Filter Preparation
    if 'record_filter' not in session:
        session['record_filter'] = {}
    if request.method == 'POST':
        search_query = request.form.get("search", session['record_filter'].get('search_query', ""))
        agree = request.form.get("agree", "")   # This is exclusively for dentist system
        filterStatus = request.form.get("filterStatus", "") 
        filterPriority = request.form.get("filterPriority", "") 
        filterProvince = request.form.get("filterProvince", "") 
        filterSpecialist = request.form.get("filterSpecialist", "")
        filterFollowup = request.form.get("filterFollowup", "")
        filterRetrain = request.form.get("filterRetrain", "")

        # Save filter parameters to the session
        session['record_filter']['search_query'] = search_query
        session['record_filter']['agree'] = agree
        session['record_filter']['filterStatus'] = filterStatus
        session['record_filter']['filterPriority'] = filterPriority
        session['record_filter']['filterProvince'] = filterProvince
        session['record_filter']['filterSpecialist'] = filterSpecialist
        session['record_filter']['filterFollowup'] = filterFollowup
        session['record_filter']['filterRetrain'] = filterRetrain

        # If there is a form submission, reset the current page to 1
        session['current_record_role'] = 1
    else:
        # Load values from session or use an empty string
        search_query = session['record_filter'].get('search_query', "")
        agree = session['record_filter'].get("agree", "")   # This is exclusively for dentist system
        filterStatus = session['record_filter'].get("filterStatus", "") 
        filterPriority = session['record_filter'].get("filterPriority", "") 
        filterProvince = session['record_filter'].get("filterProvince", "") 
        filterSpecialist = session['record_filter'].get("filterSpecialist", "")
        filterFollowup = session['record_filter'].get("filterFollowup", "")
        filterRetrain = session['record_filter'].get("filterRetrain", "")

    if 'current_record_role' in session and session['current_record_role'] == role:
        page = request.args.get("page", session['current_record_page'], type=int)
        session['current_record_page'] = page
    else: # There is a switch of role (e.g. dentist -> admin)
        page = 1 
        session['current_record_page'] = 1  # Reset currrent page to 1
        session['current_record_role'] = role # Reassign role

    session['records_per_page'] = 12 # Subject to change in the future

    if role=='specialist':
        paginated_data, supplemental_data, dataCount = record_specialist()
    elif role=='admin':
        paginated_data, supplemental_data, dataCount = record_specialist(admin=True) 
    elif role=='dentist':
        paginated_data, supplemental_data, dataCount = record_dentist()
    elif role=='osm':
        paginated_data, supplemental_data, dataCount = record_osm()
    
    # Further process each item in paginated_data
    for item in paginated_data:
        item["formatted_created_at"] = item["created_at"].strftime("%d/%m/%Y %H:%M")

    if dataCount is not None:
        dataCount = dataCount['full_count']
    else:
        dataCount = 0
    total_pages = (dataCount - 1) // session['records_per_page'] + 1

    if role == 'osm':
        load_osm_group_info()

    return render_template(
                role + "_record.html",
                dataCount=dataCount,
                paginated_data=paginated_data,
                current_page=page,
                total_pages=total_pages,
                data=supplemental_data)

# region /edit/<role>
@bp.route('/edit/<role>', methods=('GET', 'POST'))
@login_required
@role_validation
def editByRole(role):
    data = {}
    thai_months = [
                    'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 
                    'พฤษภาคม', 'มิถุนายน', 'กรกฎาคม', 
                    'สิงหาคม', 'กันยายน', 'ตุลาคม', 
                    'พฤศจิกายน', 'ธันวาคม'
                ]
    db, cursor = get_db()
    target_template = '/newTemplate/patient_edit.html'
    if role == 'patient':
        # Handling GET request to render the edit form
        if request.method == 'GET':
            sql = '''SELECT 
                        name, 
                        surname, 
                        birthdate, 
                        province, 
                        national_id, 
                        job_position, 
                        default_location, 
                        email,
                        address,
                        phone
                    FROM 
                        user
                    WHERE 
                        id = %s;
                    '''
            val = (session["user_id"],)
            cursor.execute(sql, val)
            data = cursor.fetchone()
            if data["birthdate"]:
                birthdate = data["birthdate"]
                data["birthdate_day"] = birthdate.day
                data["birthdate_month"] = thai_months[birthdate.month - 1]
                data["birthdate_year"] = birthdate.year + 543
                del data["birthdate"]
            return render_template(target_template, data=data)
        
        # Handling POST request to update data
        elif request.method == 'POST':
            data['name'] = remove_prefix(request.form['name'])
            data['surname'] = request.form['surname']
            data['national_id'] = request.form['national_id']
            data['job_position'] = request.form['job_position']
            data['address'] = request.form['address']
            data['province'] = request.form['province']
            data['email'] = request.form['email']
            data['phone'] = request.form['phone']
            data['dob_day'] = int(request.form['dob_day'])
            data['dob_month'] = int(request.form['dob_month'])
            data['dob_year'] = int(request.form['dob_year'])
            birthdate = date(data["dob_year"] - 543, data["dob_month"], data["dob_day"])
            data["valid_phone"] = True
            data["valid_province_name"] = True

            if data["email"]=='':
                data["email"] = None
            if data["phone"]=='':
                data["phone"] = None
            inval = []
            valid_func_list = [ validate_province_name,
                                validate_phone]
            for valid_func in valid_func_list:
                args = {'data': data, 'form': request.form , 'invalid':inval}
                valid_check, data , inval = valid_func(args)
                if not valid_check:
                    return render_template(target_template,data=data)


            sql = '''UPDATE user
                     SET name = %s, surname = %s, national_id = %s, job_position = %s,
                         address = %s, province = %s, email = %s, phone = %s, birthdate = %s
                     WHERE id = %s;'''
            val = (data['name'], data['surname'], data['national_id'],  data['job_position'], data['address']
                   , data['province'], data['email'], data['phone'], birthdate, session["user_id"])
            cursor.execute(sql, val)
            db.commit()
            flash('ข้อมูลส่วนตัวได้รับการแก้ไขแล้ว', 'success')
            return redirect('/edit/patient')
    elif role == 'osm':
        target_template = '/newTemplate/osm_edit.html'
        if request.method == 'GET':
            sql = '''SELECT 
                        name, 
                        surname, 
                        province,
                        hospital,
                        national_id, 
                        job_position, 
                        phone,
                        osm_job
                    FROM 
                        user
                    WHERE 
                        id = %s;
                    '''
            val = (session["user_id"],)
            cursor.execute(sql, val)
            data = cursor.fetchone()
            
            return render_template(target_template,data=data)
        if request.method == 'POST':
            data['name'] = request.form['name']
            data['surname'] = request.form['surname']
            data['job_position'] = request.form['job_position']
            data['osm_job'] = request.form.get('osm_job')
            data['license'] = request.form.get('license')
            data['hospital'] = request.form['hospital']
            data['province'] = request.form['province']
            data['national_id'] = request.form['national_id']
            data['phone'] = request.form['phone']
            data["valid_phone"] = True
            data["valid_province_name"] = True
            data["valid_national_id"] = True

            inval = []
            valid_func_list = [ validate_national_id,
                               validate_phone,
                               validate_license,
                               validate_province_name,]
            for valid_func in valid_func_list:
                args = {'data': data, 'form': request.form , 'invalid':inval}
                valid_check, data , inval = valid_func(args)
                if not valid_check:
                    return render_template(target_template,data=data)
                
            if data["osm_job"] == '':
                data["osm_job"] = None
            if data["license"] == '':
                data["license"] = None

            sql = '''UPDATE user SET 
                    name = %s,
                    surname = %s,
                    job_position = %s,
                    osm_job = %s,
                    license = %s,
                    hospital = %s,
                    province = %s,
                    national_id = %s,
                    phone = %s
                 WHERE 
                    id = %s;
                '''
            values = (data['name'], data['surname'], data['job_position'], data['osm_job'], data['license'], data['hospital']
                      , data['province'],data['national_id'],data['phone'], session["user_id"])
            cursor.execute(sql, values)
            db.commit()
            flash('ข้อมูลส่วนตัวได้รับการแก้ไขแล้ว', 'success')
            return redirect("/edit/osm")
    else:
        target_template = "/newTemplate/dentist_edit.html"
        if request.method == 'GET':
            sql = '''SELECT 
                        name, 
                        surname, 
                        province,
                        hospital,
                        national_id, 
                        job_position, 
                        phone,
                        email
                    FROM 
                        user
                    WHERE 
                        id = %s;
                    '''
            val = (session["user_id"],)
            cursor.execute(sql, val)
            data = cursor.fetchone()
            data["email"] = data["email"] if data.get("email") else ""
            return render_template(target_template,data=data)
        if request.method == "POST":
            data['name'] = request.form['name']
            data['surname'] = request.form['surname']
            data['job_position'] = request.form['job_position']
            data['osm_job'] = request.form.get('osm_job', '')
            data['license'] = request.form.get('license', '')
            data['hospital'] = request.form['hospital']
            data['province'] = request.form['province']
            data['phone'] = request.form['phone']
            data['email'] = request.form['email']
            data["valid_phone"] = True
            data["valid_province_name"] = True

            if data["email"]=='':
                data["email"] = None
            if data["phone"]=='':
                data["phone"] = None
            if data["osm_job"] == '':
                data["osm_job"] = None

            inval = []
            valid_func_list = [ validate_province_name,
                                validate_phone,
                                validate_license]
            for valid_func in valid_func_list:
                args = {'data': data, 'form': request.form , 'invalid':inval}
                valid_check, data , inval = valid_func(args)
                if not valid_check:
                    return render_template(target_template,data=data)
            
            if data["license"] == '':
                data["license"] = None
                
            sql = '''UPDATE user SET 
                    name = %s,
                    surname = %s,
                    job_position = %s,
                    osm_job = %s,
                    license = %s,
                    hospital = %s,
                    province = %s,
                    phone = %s,
                    email = %s
                 WHERE 
                    id = %s;
                '''
            values = (data['name'], data['surname'], data['job_position'], data['osm_job'], data['license'], data['hospital']
                      , data['province'], data['phone'], data['email'], session["user_id"])
            cursor.execute(sql, values)
            db.commit()
            flash('ข้อมูลส่วนตัวได้รับการแก้ไขแล้ว', 'success')
            return redirect('/edit/dentist')

# region report
@bp.route('/admin/report/')
@login_required
@admin_only
def report():
    return render_template("/newTemplate/submission_report.html")

# region admin_page
@bp.route('/admin_page/', methods=('GET','POST','DELETE'))
@login_required
@admin_only
def userManagement():
    return render_template("/newTemplate/admin_management.html")

@bp.route('/admin_record2/')
def adminRecord2():
    return render_template("/newTemplate/admin_diagnosis.html")

#followup for dentist page
@bp.route('/followup/admin', methods=('GET','POST'))
@login_required
@admin_only
def followupManage():
    page = request.args.get("page", session.get('current_record_page', 1), type=int)
    session['current_record_page'] = page
    session['records_per_page'] = 12
    paginated_data, dataCount = record_followup() 

    if dataCount is not None:
        dataCount = dataCount['total_records']
    else:
        dataCount = 0
    total_pages = (dataCount - 1) // session['records_per_page'] + 1

    return render_template("/newTemplate/admin_followup.html",dataCount=dataCount,
                current_page=page,
                total_pages=total_pages,
                data=paginated_data
                )

@bp.route('/followup/confirm/<int:submission_id>', methods=['POST'])
@login_required
@admin_only
def confirmFeedback(submission_id):
    feedback = request.form.get('feedback')
    note = request.form.get('note')

    if feedback and note:
        db, cursor = get_db()
        sql = '''
            UPDATE followup_request 
            SET followup_request_status = %s,
                followup_note = %s,
                followup_feedback = %s
            WHERE submission_id = %s;
        '''
        values = ("On contact", note, feedback, submission_id)
        cursor.execute(sql, values)
        
        # Fetch updated data for frontend update
        updated_item = {
            "id": submission_id,
            "feedback": feedback,
            "note": note
        }
        return jsonify({"status": "success", "message": "✅ ยืนยัน Feedback สำเร็จ", "item": updated_item})
    
    return jsonify({"status": "error", "message": "⚠️ กรุณาใส่ Feedback และ Note "})

@bp.route('/followup/export', methods=['GET'])
@login_required
def export_followup_records():
    db, cursor = get_db()
    # Fetch raw data from database
    sql_query = '''
        SELECT 
            sr.id,
            sr.sender_id AS "senderId", 
            pcid.case_id AS "Case ID",
            sr.fname AS "imageName"
        FROM submission_record AS sr 
        INNER JOIN followup_request AS fr ON sr.id = fr.submission_id
        LEFT JOIN patient_case_id AS pcid ON pcid.id = fr.submission_id
        ORDER BY pcid.case_id ASC
    '''
    cursor.execute(sql_query)
    records = cursor.fetchall()

    if not records:
        return "No data available", 204

    # Convert to DataFrame for processing
    df = pd.DataFrame(records)

    # Add Image URL Column (If needed)
    df["Image Link"] = "https://icohold.anamai.moph.go.th:85/load_image/upload/" + df["senderId"].astype(str) + "/" + df["imageName"].astype(str) 

    df.insert(0, "No.", range(1, len(df) + 1))
    df["Specialist Feedback"] = None  # or you can use an empty string `""`
    df["Note"] = None  # or you can use an empty string `""`
    df = df[["No.", "Case ID", "Image Link", "Specialist Feedback", "Note"]]

    # Start creating Excel file
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Followup Records"

    # Define Headers
    headers = df.columns.tolist()
    ws.append(headers)

    column_widths = {
    "No.": 10,
    "Case ID": 10,
    "Image Link": 80,
    "Specialist Feedback": 50,
    "Note": 30
    }

    # Style Headers
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold_font = Font(bold=True)

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"].fill = header_fill
        ws[f"{col_letter}1"].font = bold_font
        if header in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[header]
        else:
            ws.column_dimensions[col_letter].width = 25

    # Add Data Rows
    for row in df.itertuples(index=False):
        ws.append(row)

    # Save to buffer
    output.seek(0)
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=followup_records.xlsx"}
    )

# region edit
@bp.route('/edit/', methods=('GET','PUT','POST'))
@login_required
@admin_only
def edit():
    return render_template("/newTemplate/admin_edit.html")

# region construct_specialist_filter_sql (and admin)
# applicable for both specialist and admin
def construct_specialist_filter_sql():

    # Retrieve filter parameters from the session
    search_query = session['record_filter'].get('search_query', "")
    filterStatus = session['record_filter'].get('filterStatus', "")
    filterPriority = session['record_filter'].get('filterPriority', "")
    filterProvince = session['record_filter'].get('filterProvince', "")
    filterSpecialist = session['record_filter'].get('filterSpecialist', "")
    filterFollowup = session['record_filter'].get("filterFollowup", "")
    filterRetrain = session['record_filter'].get("filterRetrain", "")

    # SQL Filter Construction
    filter_query_list = [] # AND criteria
    search_query_list = [] # OR criteria

    if filterStatus=="1":
        filter_query_list.append('(dentist_feedback_code IS NOT NULL)')
    elif filterStatus=="0":
        filter_query_list.append('(dentist_feedback_code IS NULL)')
    if filterPriority!="" and filterPriority.isdigit():
        filter_query_list.append(f"(special_request = {int(filterPriority)})")
    if filterProvince!="":
        filter_query_list.append(f"(location_province = '{filterProvince}')")
    if filterSpecialist!="" and filterSpecialist.isdigit():
        filter_query_list.append(f"(dentist_id = {int(filterSpecialist)})")
    if filterFollowup!="":
        filter_query_list.append(f"(followup_request_status = '{filterFollowup}')")
    if filterRetrain!="":
        filter_query_list.append(f"(retrain_request_status = '{filterRetrain}')")

    filter_query = " AND ".join(filter_query_list)

    if search_query!="":
        search_query_list.append(f"(INSTR(LOWER(fname), '{search_query}'))")
        search_query_list.append(f"(patient_user.name IS NOT NULL AND (INSTR('{search_query}', patient_user.name) OR INSTR('{search_query}', patient_user.surname)))")
        search_query_list.append(f"(case_report IS NOT NULL AND (INSTR(case_report, '{search_query}')))")
        search_query_list.append(f"(dentist_feedback_code IS NOT NULL AND (LOWER(dentist_feedback_code) = '{search_query.lower()}'))")
        if search_query.isdigit():
            search_query_list.append(f"(case_id = {int(search_query)})")
        search_query_list.append(f"(location_district IS NOT NULL AND (INSTR(location_district, '{search_query}')))")
        search_query_list.append(f"(location_amphoe IS NOT NULL AND (INSTR(location_amphoe, '{search_query}')))")
        search_query_list.append(f"(location_province IS NOT NULL AND (INSTR(location_province, '{search_query}')))")
        if search_query.isdigit():
            search_query_list.append(f"(location_zipcode IS NOT NULL AND (location_zipcode = {int(search_query)}))")

        if search_query.lower() == 'opmd':
            search_query_list.append(f"(ai_prediction = 1)")
        if search_query.lower() == 'oscc':
            search_query_list.append(f"(ai_prediction = 2)")
        
        search_query_combined = " OR ".join(search_query_list)
        
        if len(filter_query_list)>0:
            filter_query = f'({filter_query}) AND ({search_query_combined})'
        else:
            filter_query = f'{search_query_combined}'
    
    supplemental_data = {}
    supplemental_data['search_query'] = search_query
    supplemental_data['filterStatus'] = filterStatus
    supplemental_data['filterPriority'] = filterPriority
    supplemental_data['filterProvince'] = filterProvince
    supplemental_data['filterSpecialist'] = filterSpecialist
    supplemental_data['filterFollowup'] = filterFollowup
    supplemental_data['filterRetrain'] = filterRetrain
    
    return filter_query, supplemental_data

# region record_specialist (and admin)
# applicable for both specialist and admin
def record_specialist(admin=False): 

    filter_query, supplemental_data = construct_specialist_filter_sql()

    # Pagination
    page = session['current_record_page']
    records_per_page = session['records_per_page']
    offset = (page-1)*records_per_page

    db, cursor = get_db()
    if admin: # Admin
        if filter_query!='':
            sql_data ='''
            SELECT
                sr.id,
                sr.case_id,
                sr.channel,
                sr.fname,
                sr.patient_name,
                sr.patient_surname,
                sr.birthdate,
                sr.sender_name,
                sr.sender_surname,
                sr.sender_hospital,
                sr.sender_id,
                sr.patient_id,
                sr.dentist_id,
                sr.sender_phone,
                sr.special_request,
                sr.location_province,
                sr.location_amphoe,
                sr.location_district,
                sr.location_zipcode,
                sr.dentist_feedback_code,
                sr.dentist_feedback_comment,
                sr.dentist_feedback_lesion,
                sr.case_report,
                sr.ai_prediction,
                sr.created_at,
                sr.retrain_request_status,
                sr.followup_request_status,
                c.full_count
            FROM (
                SELECT
                    submission_record.id,
                    patient_case_id.case_id,
                    submission_record.channel,
                    submission_record.fname,
                    patient_user.name AS patient_name,
                    patient_user.surname AS patient_surname,
                    patient_user.birthdate,
                    sender.name AS sender_name,
                    sender.surname AS sender_surname,
                    sender.hospital AS sender_hospital,
                    submission_record.sender_id,
                    submission_record.patient_id,
                    submission_record.dentist_id,
                    submission_record.sender_phone,
                    submission_record.special_request,
                    submission_record.location_province,
                    submission_record.location_amphoe,
                    submission_record.location_district,
                    submission_record.location_zipcode,
                    submission_record.dentist_feedback_code,
                    submission_record.dentist_feedback_comment,
                    submission_record.dentist_feedback_lesion,
                    submission_record.case_report,
                    submission_record.ai_prediction,
                    submission_record.created_at,
                    retrain_request.retrain_request_status,
                    followup_request.followup_request_status
                FROM submission_record
                LEFT JOIN patient_case_id ON submission_record.id = patient_case_id.id
                LEFT JOIN user AS patient_user ON submission_record.patient_id = patient_user.id
                LEFT JOIN user AS sender ON submission_record.sender_id = sender.id
                LEFT JOIN followup_request ON submission_record.id = followup_request.submission_id
                LEFT JOIN retrain_request ON submission_record.id = retrain_request.submission_id
                WHERE ''' + filter_query + '''
                ORDER BY submission_record.created_at DESC
                LIMIT %s OFFSET %s
            ) AS sr
            CROSS JOIN (
                SELECT COUNT(submission_record.id) AS full_count
                FROM submission_record
                LEFT JOIN patient_case_id ON submission_record.id = patient_case_id.id
                LEFT JOIN user AS patient_user ON submission_record.patient_id = patient_user.id
                LEFT JOIN user AS sender ON submission_record.sender_id = sender.id
                LEFT JOIN followup_request ON submission_record.id = followup_request.submission_id
                LEFT JOIN retrain_request ON submission_record.id = retrain_request.submission_id
                WHERE ''' + filter_query + '''
            ) AS c
            ORDER BY sr.created_at DESC;
            '''
            val = (records_per_page, offset)
        else:
            sql_data ='''
                SELECT
                    sr.id,
                    patient_case_id.case_id,
                    sr.channel,
                    sr.fname,
                    patient_user.name    AS patient_name,
                    patient_user.surname AS patient_surname,
                    patient_user.birthdate,
                    sender.name AS sender_name,
                    sender.surname AS sender_surname,
                    sender.hospital AS sender_hospital,
                    sr.sender_id,
                    sr.patient_id,
                    sr.dentist_id,
                    sr.sender_phone,
                    sr.special_request,
                    sr.location_province,
                    sr.location_amphoe,
                    sr.location_district,
                    sr.location_zipcode,
                    sr.dentist_feedback_comment,
                    sr.dentist_feedback_code,
                    sr.case_report,
                    sr.ai_prediction,
                    sr.created_at,
                    retrain_request.retrain_request_status,
                    followup_request.followup_request_status,
                    (SELECT COUNT(*) FROM submission_record) AS full_count
                FROM (
                    SELECT submission_record.id
                    FROM submission_record
                    ORDER BY submission_record.created_at DESC
                    LIMIT %s OFFSET %s
                ) submission_record_limited
                LEFT JOIN submission_record sr ON submission_record_limited.id = sr.id
                LEFT JOIN patient_case_id ON sr.id = patient_case_id.id
                LEFT JOIN user AS patient_user ON sr.patient_id = patient_user.id
                LEFT JOIN user AS sender ON sr.sender_id = sender.id
                LEFT JOIN followup_request ON sr.id = followup_request.submission_id
                LEFT JOIN retrain_request ON sr.id = retrain_request.submission_id
                ORDER BY sr.created_at DESC;
                '''
            val = (records_per_page, offset)
    else: # Specialist
        if filter_query!='':
            sql_data ='''
                SELECT 
                    sr.id, 
                    sr.case_id, 
                    sr.channel, 
                    sr.fname, 
                    sr.patient_name, 
                    sr.patient_surname, 
                    sr.birthdate,
                    sr.sender_id, 
                    sr.patient_id, 
                    sr.dentist_id, 
                    sr.sender_phone, 
                    sr.special_request,
                    sr.location_province, 
                    sr.location_amphoe, 
                    sr.location_district, 
                    sr.location_zipcode, 
                    sr.dentist_feedback_comment,
                    sr.dentist_feedback_code,
                    sr.case_report,
                    sr.ai_prediction, 
                    sr.created_at, 
                    sr.retrain_request_status, 
                    sr.followup_request_status,
                    c.full_count
                FROM (
                    SELECT
                        submission_record.id,
                        patient_case_id.case_id,
                        submission_record.channel,
                        submission_record.fname,
                        patient_user.name AS patient_name,
                        patient_user.surname AS patient_surname,
                        patient_user.birthdate,
                        submission_record.sender_id,
                        submission_record.patient_id,
                        submission_record.dentist_id,
                        submission_record.sender_phone,
                        submission_record.special_request,
                        submission_record.location_province,
                        submission_record.location_amphoe,
                        submission_record.location_district,
                        submission_record.location_zipcode,
                        submission_record.dentist_feedback_comment,
                        submission_record.dentist_feedback_code,
                        submission_record.case_report,
                        submission_record.ai_prediction,
                        submission_record.created_at,
                        retrain_request.retrain_request_status,
                        followup_request.followup_request_status
                    FROM submission_record
                    INNER JOIN patient_case_id 
                        ON submission_record.id = patient_case_id.id
                    LEFT JOIN user AS patient_user 
                        ON submission_record.patient_id = patient_user.id
                    LEFT JOIN followup_request 
                        ON submission_record.id = followup_request.submission_id
                    LEFT JOIN retrain_request 
                        ON submission_record.id = retrain_request.submission_id
                    WHERE ''' + filter_query + '''
                    ORDER BY submission_record.created_at DESC
                    LIMIT %s OFFSET %s
                ) AS sr
                CROSS JOIN (
                    SELECT 
                        COUNT(submission_record.id) AS full_count
                    FROM submission_record
                    INNER JOIN patient_case_id 
                        ON submission_record.id = patient_case_id.id
                    LEFT JOIN user AS patient_user 
                        ON submission_record.patient_id = patient_user.id
                    LEFT JOIN followup_request 
                        ON submission_record.id = followup_request.submission_id
                    LEFT JOIN retrain_request 
                        ON submission_record.id = retrain_request.submission_id
                    WHERE ''' + filter_query + '''
                ) AS c
                ORDER BY sr.created_at DESC;
                '''
            val = (records_per_page, offset)
        else:
            sql_data ='''
                SELECT 
                    sr.id, 
                    submission_record_limited.case_id, 
                    sr.channel, 
                    sr.fname, 
                    patient_user.name as patient_name, 
                    patient_user.surname as patient_surname, 
                    patient_user.birthdate,
                    sr.sender_id, 
                    sr.patient_id, 
                    sr.dentist_id, 
                    sr.sender_phone, 
                    sr.special_request,
                    sr.location_province, 
                    sr.location_amphoe, 
                    sr.location_district, 
                    sr.location_zipcode, 
                    sr.dentist_feedback_comment,
                    sr.dentist_feedback_code,
                    sr.case_report,
                    sr.ai_prediction, 
                    sr.created_at, 
                    retrain_request.retrain_request_status, 
                    followup_request.followup_request_status,
                    c.full_count
                FROM (
                    SELECT 
                        submission_record.id, 
                        patient_case_id.case_id
                    FROM submission_record
                    INNER JOIN patient_case_id 
                        ON submission_record.id = patient_case_id.id
                    ORDER BY submission_record.created_at DESC
                    LIMIT %s OFFSET %s
                ) AS submission_record_limited
                LEFT JOIN submission_record AS sr 
                    ON submission_record_limited.id = sr.id
                LEFT JOIN user AS patient_user 
                    ON sr.patient_id = patient_user.id
                LEFT JOIN followup_request 
                    ON sr.id = followup_request.submission_id
                LEFT JOIN retrain_request 
                    ON sr.id = retrain_request.submission_id
                CROSS JOIN (
                    SELECT 
                        COUNT(submission_record.id) AS full_count
                    FROM submission_record
                    INNER JOIN patient_case_id 
                        ON submission_record.id = patient_case_id.id
                ) AS c
                ORDER BY sr.created_at DESC;
                '''
            val = (records_per_page, offset)
    
    cursor.execute(sql_data,val)
    paginated_data = cursor.fetchall()
    if len(paginated_data)>0:
        dataCount = {'full_count': paginated_data[0]['full_count']}
    else:
        dataCount = {'full_count': 0}

    # Process each item in paginated_data
    for item in paginated_data:
        if item['channel']=='PATIENT':
            item['owner_id'] = item['patient_id']
        else:
            item['owner_id'] = item['sender_id']
        if ("birthdate" in item and item["birthdate"]):
            item["age"] = calculate_age(item["birthdate"])

    if admin:
        sql = '''SELECT location_province
            FROM submission_record
            GROUP BY location_province
            ORDER BY COUNT(location_province) DESC;'''
    else:
        sql = '''SELECT location_province
            FROM submission_record
            WHERE channel != 'DENTIST'
            GROUP BY location_province
            ORDER BY COUNT(location_province) DESC;'''
    cursor.execute(sql)
    dictList = cursor.fetchall()
    supplemental_data['province_name_list'] = [item['location_province'] for item in dictList]
    
    sql = '''SELECT name, surname, license, user.id
        FROM submission_record
        LEFT JOIN user ON submission_record.dentist_id=user.id
        WHERE channel != 'DENTIST' AND submission_record.dentist_id IS NOT NULL
        GROUP BY user.id
        ORDER BY COUNT(user.id) DESC'''
    cursor.execute(sql)
    dictList = cursor.fetchall()
    supplemental_data['specialist_list'] = []
    for item in dictList:
        if item['license'] is not None:
            supplemental_data['specialist_list'].append((item['id'], f"{item['name']} {item['surname']} ({item['license']})"))
        else:
            supplemental_data['specialist_list'].append((item['id'], f"{item['name']} {item['surname']}"))
    
    return paginated_data, supplemental_data, dataCount

# region construct_dentist_filter_sql
def construct_dentist_filter_sql():

    # Retrieve filter parameters from the session
    search_query = session['record_filter'].get('search_query', "")
    agree = session['record_filter'].get('agree', "")

    # SQL Filter Construction
    search_query_list = [] # OR criteria

    if search_query!="":
        search_query_list.append(f"(INSTR(LOWER(fname), '{search_query}'))")
        search_query_list.append(f"(dentist_feedback_comment IS NOT NULL AND (INSTR(dentist_feedback_comment, '{search_query}')))")
        search_query_list.append(f"(dentist_feedback_code IS NOT NULL AND (LOWER(dentist_feedback_code) = '{search_query.lower()}'))")
        search_query_list.append(f"(location_district IS NOT NULL AND (INSTR(location_district, '{search_query}')))")
        search_query_list.append(f"(location_amphoe IS NOT NULL AND (INSTR(location_amphoe, '{search_query}')))")
        search_query_list.append(f"(location_province IS NOT NULL AND (INSTR(location_province, '{search_query}')))")
        if search_query.isdigit():
            search_query_list.append(f"(location_zipcode IS NOT NULL AND (location_zipcode = {int(search_query)}))")

        if search_query.lower() == 'opmd':
            search_query_list.append(f"(ai_prediction = 1)")
        if search_query.lower() == 'oscc':
            search_query_list.append(f"(ai_prediction = 2)")
        
        search_query_combined = " OR ".join(search_query_list)
        
        filter_query = f'({search_query_combined})'
    else:
        filter_query = ''
    
    supplemental_data = {}
    supplemental_data['search_query'] = search_query
    supplemental_data['agree'] = agree

    return filter_query, supplemental_data

# region record_dentist
def record_dentist(): 

    filter_query, supplemental_data = construct_dentist_filter_sql()

    # Pagination
    page = session['current_record_page']
    records_per_page = session['records_per_page']
    offset = (page-1)*records_per_page

    db, cursor = get_db()
    if filter_query!='':
        sql_data ='''
            SELECT 
                sr.id,
                sr.channel,
                sr.fname,
                sr.sender_id,
                sr.dentist_feedback_code,
                sr.dentist_feedback_comment,
                sr.ai_prediction,
                sr.location_district,
                sr.location_amphoe,
                sr.location_province,
                sr.location_zipcode,
                sr.created_at,
                c.full_count
            FROM (
                SELECT 
                    submission_record.id,
                    submission_record.channel,
                    submission_record.fname,
                    submission_record.sender_id,
                    submission_record.dentist_feedback_code,
                    submission_record.dentist_feedback_comment,
                    submission_record.ai_prediction,
                    submission_record.location_district,
                    submission_record.location_amphoe,
                    submission_record.location_province,
                    submission_record.location_zipcode,
                    submission_record.created_at
                FROM submission_record
                WHERE 
                    (channel = 'DENTIST' AND sender_id = %s) AND ''' + filter_query + '''
                ORDER BY submission_record.created_at DESC
                LIMIT %s OFFSET %s
            ) AS sr
            CROSS JOIN (
                SELECT 
                    COUNT(submission_record.id) AS full_count
                FROM submission_record
                WHERE 
                    (channel = 'DENTIST' AND sender_id = %s) AND ''' + filter_query + '''
            ) AS c;
            '''
        val = (session['user_id'], records_per_page, offset, session['user_id'])
    else:
        sql_data ='''
            SELECT 
                sr.id,
                sr.channel,
                sr.fname,
                sr.sender_id,
                sr.dentist_feedback_code,
                sr.dentist_feedback_comment,
                sr.ai_prediction,
                sr.location_district,
                sr.location_amphoe,
                sr.location_province,
                sr.location_zipcode,
                sr.created_at,
                c.full_count
            FROM (
                SELECT 
                    submission_record.id,
                    submission_record.channel,
                    submission_record.fname,
                    submission_record.sender_id,
                    submission_record.dentist_feedback_code,
                    submission_record.dentist_feedback_comment,
                    submission_record.ai_prediction,
                    submission_record.location_district,
                    submission_record.location_amphoe,
                    submission_record.location_province,
                    submission_record.location_zipcode,
                    submission_record.created_at
                FROM submission_record
                WHERE (channel = 'DENTIST' AND sender_id = %s)
                ORDER BY submission_record.created_at DESC
                LIMIT %s OFFSET %s
            ) AS sr
            CROSS JOIN (
                SELECT 
                    COUNT(submission_record.id) AS full_count
                FROM submission_record
                WHERE (channel = 'DENTIST' AND sender_id = %s)
            ) AS c;
            ORDER BY sr.created_at DESC
            '''
        val = (session['user_id'], records_per_page, offset, session['user_id'])

    cursor.execute(sql_data,val)
    paginated_data = cursor.fetchall()
    if len(paginated_data)>0:
        dataCount = {'full_count': paginated_data[0]['full_count']}
    else:
        dataCount = {'full_count': 0}

    return paginated_data, supplemental_data, dataCount

# region construct_osm_filter_sql
def construct_osm_filter_sql():
    
    # Retrieve filter parameters from the session
    search_query = session['record_filter'].get('search_query', "")
    filterStatus = session['record_filter'].get('filterStatus', "")
    filterPriority = session['record_filter'].get('filterPriority', "")

    # SQL Filter Construction
    filter_query_list = [] # AND criteria
    search_query_list = [] # OR criteria

    if filterStatus=="1":
        filter_query_list.append('(dentist_feedback_code IS NOT NULL)')
    elif filterStatus=="0":
        filter_query_list.append('(dentist_feedback_code IS NULL)')
    if filterPriority!="" and filterPriority.isdigit():
        filter_query_list.append(f"(special_request = {int(filterPriority)})")

    filter_query = " AND ".join(filter_query_list)

    if search_query!="":
        search_query_list.append(f"(INSTR(LOWER(fname), '{search_query}'))")
        search_query_list.append(f"(patient_user.name IS NOT NULL AND (INSTR('{search_query}', patient_user.name) OR INSTR('{search_query}', patient_user.surname)))")
        search_query_list.append(f"(dentist_feedback_code IS NOT NULL AND (LOWER(dentist_feedback_code) = '{search_query.lower()}'))")
        if search_query.isdigit():
            search_query_list.append(f"(case_id = {int(search_query)})")
        search_query_list.append(f"(location_district IS NOT NULL AND (INSTR(location_district, '{search_query}')))")
        search_query_list.append(f"(location_amphoe IS NOT NULL AND (INSTR(location_amphoe, '{search_query}')))")
        search_query_list.append(f"(location_province IS NOT NULL AND (INSTR(location_province, '{search_query}')))")
        if search_query.isdigit():
            search_query_list.append(f"(location_zipcode IS NOT NULL AND (location_zipcode = {int(search_query)}))")

        if search_query.lower() == 'opmd':
            search_query_list.append(f"(ai_prediction = 1)")
        if search_query.lower() == 'oscc':
            search_query_list.append(f"(ai_prediction = 2)")
        
        search_query_combined = " OR ".join(search_query_list)
        
        if len(filter_query_list)>0:
            filter_query = f'({filter_query}) AND ({search_query_combined})'
        else:
            filter_query = f'({search_query_combined})'
    
    supplemental_data = {}
    supplemental_data['search_query'] = search_query
    supplemental_data['filterStatus'] = filterStatus
    supplemental_data['filterPriority'] = filterPriority

    return filter_query, supplemental_data

# region construct_osm_filter_sql
def record_osm():
    
    filter_query, supplemental_data = construct_osm_filter_sql()

    # Pagination
    page = session['current_record_page']
    records_per_page = session['records_per_page']
    offset = (page-1)*records_per_page

    db, cursor = get_db()
    if filter_query!='':
        sql_data ='''
            SELECT 
                sr.id,
                sr.channel,
                sr.fname,
                sr.patient_name,
                sr.patient_surname,
                sr.birthdate,
                sr.province,
                sr.case_id,
                sr.sender_id,
                sr.patient_id,
                sr.dentist_id,
                sr.special_request,
                sr.sender_phone,
                sr.location_province,
                sr.location_zipcode,
                sr.dentist_feedback_comment,
                sr.dentist_feedback_code,
                sr.ai_prediction,
                sr.created_at,
                c.full_count
            FROM (
                SELECT 
                    submission_record.id,
                    submission_record.channel,
                    submission_record.fname,
                    patient_user.name AS patient_name,
                    patient_user.surname AS patient_surname,
                    patient_user.birthdate,
                    patient_user.province,
                    patient_case_id.case_id,
                    submission_record.sender_id,
                    submission_record.patient_id,
                    submission_record.dentist_id,
                    submission_record.special_request,
                    submission_record.sender_phone,
                    submission_record.location_province,
                    submission_record.location_zipcode,
                    submission_record.dentist_feedback_comment,
                    submission_record.dentist_feedback_code,
                    submission_record.ai_prediction,
                    submission_record.created_at
                FROM (
                    SELECT id
                    FROM submission_record
                    WHERE
                        (channel = 'OSM' AND sender_id = %s)
                        OR
                        (channel = 'PATIENT' AND submission_record.sender_phone = %s)
                ) AS submission_record_limited
                INNER JOIN submission_record
                    ON submission_record.id = submission_record_limited.id
                LEFT JOIN patient_case_id
                    ON submission_record.id = patient_case_id.id
                LEFT JOIN user AS patient_user
                    ON submission_record.patient_id = patient_user.id
                LEFT JOIN user AS sender_user
                    ON submission_record.sender_phone = sender_user.phone
                WHERE ''' + filter_query + '''
                ORDER BY submission_record.created_at DESC
                LIMIT %s OFFSET %s
            ) AS sr
            CROSS JOIN (
                SELECT 
                    COUNT(submission_record.id) AS full_count
                FROM (
                    SELECT id
                    FROM submission_record
                    WHERE
                        (channel = 'OSM' AND sender_id = %s)
                        OR
                        (channel = 'PATIENT' AND submission_record.sender_phone = %s)
                ) AS submission_record_limited
                INNER JOIN submission_record
                    ON submission_record.id = submission_record_limited.id
                LEFT JOIN patient_case_id
                    ON submission_record.id = patient_case_id.id
                LEFT JOIN user AS patient_user
                    ON submission_record.patient_id = patient_user.id
                LEFT JOIN user AS sender_user
                    ON submission_record.sender_phone = sender_user.phone
                WHERE ''' + filter_query + '''
            ) AS c
            ORDER BY sr.created_at DESC;
            '''
        val = (session['user_id'],g.user['phone'],records_per_page, offset,session['user_id'],g.user['phone'])
    else:
        sql_data ='''
            SELECT 
                sr.id,
                sr.channel,
                sr.fname,
                sr.patient_name,
                sr.patient_surname,
                sr.birthdate,
                sr.province,
                sr.case_id,
                sr.sender_id,
                sr.patient_id,
                sr.dentist_id,
                sr.special_request,
                sr.sender_phone,
                sr.location_province,
                sr.location_zipcode,
                sr.dentist_feedback_comment,
                sr.dentist_feedback_code,
                sr.ai_prediction,
                sr.created_at,
                c.full_count
            FROM (
                SELECT 
                    submission_record.id,
                    submission_record.channel,
                    submission_record.fname,
                    patient_user.name AS patient_name,
                    patient_user.surname AS patient_surname,
                    patient_user.birthdate,
                    patient_user.province,
                    patient_case_id.case_id,
                    submission_record.sender_id,
                    submission_record.patient_id,
                    submission_record.dentist_id,
                    submission_record.special_request,
                    submission_record.sender_phone,
                    submission_record.location_province,
                    submission_record.location_zipcode,
                    submission_record.dentist_feedback_comment,
                    submission_record.dentist_feedback_code,
                    submission_record.ai_prediction,
                    submission_record.created_at
                FROM (
                    SELECT id
                    FROM submission_record
                    WHERE
                        (channel = 'OSM' AND sender_id = %s)
                        OR
                        (channel = 'PATIENT' AND submission_record.sender_phone = %s)
                    ORDER BY submission_record.created_at DESC
                    LIMIT %s OFFSET %s
                ) AS submission_record_limited
                INNER JOIN submission_record
                    ON submission_record.id = submission_record_limited.id
                LEFT JOIN patient_case_id
                    ON submission_record.id = patient_case_id.id
                LEFT JOIN user AS patient_user
                    ON submission_record.patient_id = patient_user.id
                LEFT JOIN user AS sender_user
                    ON submission_record.sender_phone = sender_user.phone
            ) AS sr
            CROSS JOIN (
                SELECT 
                    COUNT(submission_record.id) AS full_count
                FROM (
                    SELECT id
                    FROM submission_record
                    WHERE
                        (channel = 'OSM' AND sender_id = %s)
                        OR
                        (channel = 'PATIENT' AND submission_record.sender_phone = %s)
                ) AS submission_record_limited
                INNER JOIN submission_record
                    ON submission_record.id = submission_record_limited.id
            ) AS c
            ORDER BY sr.created_at DESC;
            '''
        val = (session['user_id'], g.user['phone'], records_per_page, offset, session['user_id'], g.user['phone'])
    
    cursor.execute(sql_data,val)
    paginated_data = cursor.fetchall()
    if len(paginated_data)>0:
        dataCount = {'full_count': paginated_data[0]['full_count']}
    else:
        dataCount = {'full_count': 0}

    # Process each item in paginated_data
    for item in paginated_data:
        if item['channel']=='PATIENT':
            item['owner_id'] = item['patient_id']
        else:
            item['owner_id'] = item['sender_id']
        if ("birthdate" in item and item["birthdate"]):
            item["age"] = calculate_age(item["birthdate"])
    
    return paginated_data, supplemental_data, dataCount

# region record_patient
def record_patient(): 
    
    db, cursor = get_db()
    sql = '''SELECT submission_record.id, case_id, special_request, sender_id, patient_id, sender_phone, channel, 
                dentist_feedback_code, dentist_feedback_comment, dentist_feedback_date, ai_prediction, ai_scores, submission_record.created_at
            FROM submission_record
            INNER JOIN patient_case_id ON submission_record.id=patient_case_id.id
            WHERE patient_id=%s
            ORDER BY case_id DESC'''
    val = (session["user_id"],)
    cursor.execute(sql, val)
    db_query = cursor.fetchall()
    data=db_query
    for item in data:
        item['thai_datetime'] = format_thai_datetime(item['created_at'])
        if item['ai_prediction']==0 and item['dentist_feedback_code']=='NORMAL':
            item['dentistCommentAgreeCode'] = 'TN'
        elif item['ai_prediction']==0 and (item['dentist_feedback_code']=='OPMD' or item['dentist_feedback_code']=='OSCC'):
            item['dentistCommentAgreeCode'] = 'FN'
        elif item['ai_prediction']!=0 and (item['dentist_feedback_code']=='OPMD' or item['dentist_feedback_code']=='OSCC'):
            item['dentistCommentAgreeCode'] = 'TP'
        elif item['ai_prediction']!=0 and item['dentist_feedback_code']=='NORMAL':
            item['dentistCommentAgreeCode'] = 'FP'
        else:
            item['dentistCommentAgreeCode'] = 'Error'
            if item['dentist_feedback_comment'] == 'NON_STANDARD':
                item['dentistComment'] = 'มุมมองไม่ได้มาตรฐาน'
            elif item['dentist_feedback_comment'] == 'BLUR':
                item['dentistComment'] = 'ภาพเบลอ ไม่ชัด'
            elif item['dentist_feedback_comment'] == 'DARK':
                item['dentistComment'] = 'ภาพช่องปากมืดเกินไป ขอเปิดแฟลชด้วย'
            elif item['dentist_feedback_comment'] == 'SMALL':
                item['dentistComment'] = 'ช่องปากเล็กเกินไป ขอนำกล้องเข้าใกล้ปากมากกว่านี้'
    return data

# region update_submission_record
# This function will be called by image.upload_submission_module
def update_submission_record(ai_predictions, ai_scores):
    # Add the prediction record to the database
    for i, filename in enumerate(session['imageNameList']):
        db, cursor = get_db()
        
        if session['login_mode']=='dentist':
            
            if g.user['default_location'] is None or str(g.user['default_location'])!=str(session['location']):
                sql = "UPDATE user SET default_location=%s WHERE id=%s"
                val = (str(session['location']), session['user_id'])
                cursor.execute(sql, val)
                reload_user_profile(session['user_id']) # Update the user info on g variable

            sql = '''INSERT INTO submission_record
                (fname,
                sender_id,
                location_district,
                location_amphoe,
                location_province,
                location_zipcode,
                ai_prediction,
                ai_scores,
                lesion_ai_version,
                quality_ai_prediction,
                quality_ai_version,
                ai_updated_at,
                channel)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'''
            val = (filename,
                    session['user_id'],
                    session['location']['district'],
                    session['location']['amphoe'],
                    session['location']['province'],
                    session['location']['zipcode'],
                    ai_predictions[i],
                    ai_scores[i],
                    current_app.config['AI_LESION_VER'],
                    session['quality_ai_predictions'][i],
                    current_app.config['AI_QUALITY_VER'],
                    datetime.now(),
                    'DENTIST')
            cursor.execute(sql, val)


        elif session['login_mode']=='patient':

            if g.user['default_location'] is None or str(g.user['default_location'])!=str(session['location']):
                sql = "UPDATE user SET default_location=%s WHERE id=%s"
                val = (str(session['location']), session['user_id'])
                cursor.execute(sql, val)
                reload_user_profile(session['user_id']) # Update the user info on g variable

            if 'sender_phone' in session and session['sender_phone']:
                sql = "UPDATE user SET default_sender_phone=%s WHERE id=%s"
                val = (session['sender_phone'], session['user_id'])
                cursor.execute(sql, val)
                reload_user_profile(session['user_id']) # Update the user info on g variable
            
            if (g.user['default_sender_phone'] and ('sender_phone' not in session or session['sender_phone'] is None)):
                sql = "UPDATE user SET default_sender_phone=NULL WHERE id=%s"
                val = (session['user_id'], )
                cursor.execute(sql, val)
                reload_user_profile(session['user_id']) # Update the user info on g variable
            
            sql = '''INSERT INTO submission_record 
                    (fname,
                    sender_id,
                    sender_phone,
                    location_district,
                    location_amphoe,
                    location_province,
                    location_zipcode,
                    patient_id,
                    ai_prediction,
                    ai_scores,
                    lesion_ai_version,
                    quality_ai_prediction,
                    quality_ai_version,
                    ai_updated_at,
                    channel)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'''
            val = (filename,
                    session['sender_id'],
                    session['sender_phone'],
                    session['location']['district'],
                    session['location']['amphoe'],
                    session['location']['province'],
                    session['location']['zipcode'],
                    session['user_id'],
                    ai_predictions[i],
                    ai_scores[i],
                    current_app.config['AI_LESION_VER'],
                    session['quality_ai_predictions'][i],
                    current_app.config['AI_QUALITY_VER'],
                    datetime.now(),
                    'PATIENT')
            cursor.execute(sql, val)
            
            cursor.execute("SELECT LAST_INSERT_ID()")
            row = cursor.fetchone()
            sql = "INSERT INTO patient_case_id (id) VALUES (%s)"
            val = (row['LAST_INSERT_ID()'],)
            cursor.execute(sql, val)

        elif session['login_mode']=='osm':

            if g.user['default_location'] is None or str(g.user['default_location'])!=str(session['location']):
                sql = "UPDATE user SET default_location=%s WHERE id=%s"
                val = (str(session['location']), session['user_id'])
                cursor.execute(sql, val)
                reload_user_profile(session['user_id']) # Update the user info on g variable

            sql = '''INSERT INTO submission_record 
                    (fname, 
                    sender_id,
                    location_district,
                    location_amphoe,
                    location_province,
                    location_zipcode,
                    patient_id,
                    patient_national_id,
                    ai_prediction,
                    ai_scores,
                    lesion_ai_version,
                    quality_ai_prediction,
                    quality_ai_version,
                    ai_updated_at,
                    channel)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'''
            val = (filename,
                    session['user_id'],
                    session['location']['district'],
                    session['location']['amphoe'],
                    session['location']['province'],
                    session['location']['zipcode'],
                    session['patient_id'],
                    session['patient_national_id'],
                    ai_predictions[i],
                    ai_scores[i],
                    current_app.config['AI_LESION_VER'],
                    session['quality_ai_predictions'][i],
                    current_app.config['AI_QUALITY_VER'],
                    datetime.now(),
                    'OSM')
            cursor.execute(sql, val)

            cursor.execute("SELECT LAST_INSERT_ID()")
            row = cursor.fetchone()
            sql = "INSERT INTO patient_case_id (id) VALUES (%s)"
            val = (row['LAST_INSERT_ID()'],)
            cursor.execute(sql, val)

def record_followup():
    
    page = session['current_record_page']
    records_per_page = session['records_per_page']
    offset = (page-1)*records_per_page
    db, cursor = get_db()
    sql_select_part = '''
                SELECT 
                sr.id, 
                sr.ai_prediction,
                sr.channel,
                sr.sender_id,
                sr.patient_id,
                sr.fname,
                pcid.case_id,
                fr.followup_request_status,
                fr.followup_note,
                fr.followup_feedback
            '''
    sql_join_part = '''
                FROM submission_record as sr 
                INNER JOIN followup_request as fr ON sr.id = fr.submission_id
                LEFT JOIN patient_case_id as pcid ON pcid.id=fr.submission_id
                '''
    sql_limit_part = '''
                ORDER BY 
                    CASE
                        WHEN fr.followup_request_status = 'initiated' THEN 0
                        ELSE 1
                    END,
                    sr.id DESC
                LIMIT %s
                OFFSET %s
            '''
    sql_count = 'SELECT Count(*) AS total_records'
                
    sql1 = sql_count + sql_join_part
    sql2 = sql_select_part + sql_join_part + sql_limit_part

    val = (records_per_page,offset)
    cursor.execute(sql1)
    dataCount = cursor.fetchone()

    cursor.execute(sql2,val)
    paginated_data = cursor.fetchall()

    for item in paginated_data:
        if item['channel']=='PATIENT':
            item['owner_id'] = item['patient_id']
        else:
            item['owner_id'] = item['sender_id']
    return paginated_data,dataCount


@bp.route('/about')
def about():
    return render_template("about.html")

@bp.route('/example')
def example():
    img_names =['oscc1.png', 'oscc2.png', 'oscc3.png', 'oscc4.png', 'oscc5.png', 'oscc6.png',
                'opmd1.png', 'opmd2.png', 'opmd3.png', 'opmd4.png', 'opmd5.png', 'opmd6.png',
                'opmd7.png', 'opmd8.png', 'opmd10.png', 'opmd11.png', 'opmd12.png', 'opmd13.png',
                'opmd14.png','opmd15.png','opmd16.png', 'opmd17.png', 'opmd18.png', 'opmd19.png', 'opmd20.png',
                'normal1.png', 'normal2.png', 'normal3.png','normal4.png', 'normal5.png',
                'normal6.png', 'normal7.png', 'normal8.png', 'normal9.png', 'normal10.png']

    texts = [
        '''ตัวอย่างที่ 1 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นมะเร็งชนิด Oral Squamous Cell Carcinoma (OSCC) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นสูงถึง 94.01%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 2 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นมะเร็งชนิด Oral Squamous Cell Carcinoma (OSCC) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นสูงถึง 96.93%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 3 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นมะเร็งชนิด Oral Squamous Cell Carcinoma (OSCC) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นสูงถึง 93.83%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
    '''ตัวอย่างที่ 4 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นมะเร็งชนิด Oral Squamous Cell Carcinoma (OSCC) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นสูงถึง 94.31%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
    '''ตัวอย่างที่ 5 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ ที่ถูกระบุว่าเป็น Oral Squamous Cell Carcinoma (OSCC) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะที่ 93.21% 
    ภาพนี้ผู้พัฒนานำมาจาก Google Search Engine'''
        ,
        '''ตัวอย่างที่ 6 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ ที่ถูกระบุว่าเป็น Oral Squamous Cell Carcinoma (OSCC) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะที่ 96.34%
    ภาพนี้ผู้พัฒนานำมาจาก Google Search Engine'''
        ,
        '''ตัวอย่างที่ 7 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 68.28% 
        โปรดสังเกตุว่า หากช่องปากปรากฎรอยโรคหลายรอยหรือรอยโรคกินเนื้อเยื่อในบริเวณกว้าง ระบบจะระบุรอยที่ชัดเจนที่สุดเท่านั้น 
        ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''
        ตัวอย่างที่ 8 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 96.05%
    โปรดสังเกตุว่า หากช่องปากปรากฎรอยโรคหลายรอยหรือรอยโรคกินเนื้อเยื่อในบริเวณกว้าง ระบบจะระบุรอยที่ชัดเจนที่สุดเท่านั้น
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 9 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 85.95%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์
    '''
        ,
        '''ตัวอย่างที่ 10 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 89.62%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 11 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 90.41%
    โปรดสังเกตุว่า หากช่องปากปรากฎรอยโรคหลายรอยหรือรอยโรคกินเนื้อเยื่อในบริเวณกว้าง ระบบจะระบุรอยที่ชัดเจนที่สุดเท่านั้น 
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 12 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 92.98%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 13 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 94.57%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 14 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 90.19%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 15 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 84.11%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 16 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 85.26%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 17 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 92.87%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 18 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 94.53%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 19 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 88.06%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 20 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 89.73%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.ดร. จิตจิโรจน์ อิทธิชัยเจริญ คณะทันตแพทยศาสตร์ มหาวิทยาลัยเชียงใหม่'''
        ,
        '''ตัวอย่างที่ 21 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ที่ได้รับการวินิจฉัยทางพยาธิวิทยาว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 89.05%
    ภาพนี้ได้รับจากการลงพื้นที่ตรวจคัดกรองของ ศทป. โดยใช้กล้องมือถือในการถ่าย'''
        ,
        '''ตัวอย่างที่ 22 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ ที่ถูกระบุว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 90.61%
        ภาพนี้ผู้พัฒนานำมาจาก Google Search Engine'''
        ,
        '''ตัวอย่างที่ 23 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ ที่ถูกระบุว่าเป็นโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMF) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 78.82%
    ภาพนี้ผู้พัฒนานำมาจาก Google Search Engine'''
        ,
        '''ตัวอย่างที่ 24 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ ที่ถูกระบุว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 82.57%
    ภาพนี้ผู้พัฒนานำมาจาก Google Search Engine'''
        ,
        '''ตัวอย่างที่ 25 รูปภาพผลลัพธ์การระบุรอยโรคของภาพถ่ายช่องปากของคนไข้ ที่ถูกระบุว่าเป็นรอยโรคก่อนมะเร็ง หรือ Oral Potentially Malignant Disorder (OPMD) โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 88.30%
    ภาพนี้ผู้พัฒนานำมาจาก Google Search Engine'''
        ,
        '''ตัวอย่างที่ 26 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 100%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 27 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 100%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 28 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 100%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 29 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 99.98%
    ภาพนี้ได้รับความอนุเคราะห์จาก ผศ.ทพ.กฤษสิทธิ์ วารินทร์ คณะทันตแพทยศาสตร์ มหาวิทยาลัยธรรมศาสตร์'''
        ,
        '''ตัวอย่างที่ 30 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 100%
    ภาพนี้ได้รับจากการลงพื้นที่ตรวจคัดกรองของ ศทป. โดยใช้กล้องมือถือในการถ่าย
    ภาพนี้มีการแก้ไขทางคอมพิวเตอร์โดยทำให้เกิดการสะท้อนกระจกซ้ายขวาเพื่อประโยชน์ต่อการแสดงผล'''
        ,
        '''ตัวอย่างที่ 31 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 99.99%
    ภาพนี้ได้รับจากการลงพื้นที่ตรวจคัดกรองของ ศทป. โดยใช้กล้องมือถือในการถ่าย
    ภาพนี้มีการแก้ไขทางคอมพิวเตอร์โดยทำให้เกิดการสะท้อนกระจกซ้ายขวาเพื่อประโยชน์ต่อการแสดงผล'''
        ,
        '''ตัวอย่างที่ 32 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 99.88% 
    ภาพนี้ได้รับจากการลงพื้นที่ตรวจคัดกรองของ ศทป. โดยใช้กล้องมือถือในการถ่าย
    ภาพนี้มีการแก้ไขทางคอมพิวเตอร์โดยทำให้เกิดการสะท้อนกระจกซ้ายขวาเพื่อประโยชน์ต่อการแสดงผล'''
        ,
        '''ตัวอย่างที่ 33 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 99.86%
    ภาพนี้ได้รับจากการลงพื้นที่ตรวจคัดกรองของ ศทป. โดยใช้กล้องมือถือในการถ่าย'''
        ,
        '''ตัวอย่างที่ 34 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 99.97%
    ภาพนี้ได้รับจากการลงพื้นที่ตรวจคัดกรองของ ศทป. โดยใช้กล้องมือถือในการถ่าย'''
        ,
        '''ตัวอย่างที่ 35 ตัวอย่างรูปภาพที่ไม่พบรอยโรค หรือ Normal โดยระบบปัญญาประดิษฐ์แจ้งค่าความน่าจะเป็นที่ 99.89%
    ภาพนี้ได้รับจากการลงพื้นที่ตรวจคัดกรองของ ศทป. โดยใช้กล้องมือถือในการถ่าย
    ภาพนี้มีการแก้ไขทางคอมพิวเตอร์โดยทำให้เกิดการสะท้อนกระจกซ้ายขวาเพื่อประโยชน์ต่อการแสดงผล'''
    ]
    data = {}
    data["imgs_name"] = img_names
    data["texts"] = texts
    return render_template("example.html", data=data)